"""
Bulk Upload Service for FormsApp

This service handles bulk upload operations for forms using Excel files.
It provides validation, error handling, and batch processing capabilities.

Features:
- Multi-sheet Excel template (Forms, Sections, Fields)
- Validates all sheets before processing
- Checks for missing required fields
- Validates dropdown selections against active database records
- Validates cross-sheet references (form_title, section_name)
- Validates JSON fields (options, dependency, additional_info)
- Prevents duplicate entries
- Provides detailed error messages with row numbers and sheet names
- Creates Forms, FormSections, and FormFields in atomic transactions

Supported file formats:
- Excel (.xlsx, .xls)

Usage:
    from ..services.bulk_upload_forms_services import bulk_upload_forms_service

    result = bulk_upload_forms_service(excel_file=file, user=request.user)
"""

import json
import logging
from openpyxl import load_workbook
from django.db import transaction

from ..models import (
    FormType, Form, FormSections, FormFields,
    FieldType, DataType, FormDraft
)

logger = logging.getLogger(__name__)


# ===============================
# Column Mapping Configuration
# ===============================

COLUMN_MAPPING = {
    # Forms sheet columns
    "Form Title": "form_title",
    "Form Type": "form_type",
    "Description": "description",
    "Is Completed": "is_completed",

    # Sections sheet columns
    "Section Name": "section_name",
    "Section Description": "section_description",
    "Section Order": "section_order",
    "Dependency Section": "dependency_section",
    "Dependency Field": "dependency_field",
    "Dependency Option": "dependency_option",
    "Dependency (JSON)": "dependency",

    # Fields sheet columns
    "Field Label": "field_label",
    "Field Type": "field_type",
    "Data Type": "data_type",
    "Required": "required",
    "Field Order": "field_order",
    "Width": "width",
    "Options": "options",
    "Validation": "validation",
    "Parent Field": "parent_field",
    "Field Dep Section": "field_dep_section",
    "Field Dep Field": "field_dep_field",
    "Field Dep Option": "field_dep_option",
    "Dependency": "field_dependency",
    "Additional Info": "additional_info",
}


# Required fields for each sheet
REQUIRED_FIELDS = {
    "forms": ["form_title", "form_type"],
    "sections": ["form_title", "section_name", "section_order"],
    "fields": ["form_title", "section_name", "field_label",
               "field_type", "data_type", "field_order"],
}


# Dropdown fields that must be selected from valid options
DROPDOWN_FIELDS = {
    "forms": {
        "form_type": "Form Type"
    },
    "fields": {
        "field_type": "Field Type",
        "data_type": "Data Type"
    }
}


# Sheet names in the Excel template
SHEET_NAMES = {
    "forms": "Forms",
    "sections": "Sections",
    "fields": "Fields",
    "references": "References"
}


# ===============================
# Main Bulk Upload Function
# ===============================

def bulk_upload_forms_service(file, user):
    """
    Main bulk upload service function for forms.

    Processes uploaded Excel file with multiple sheets and creates forms in bulk.
    Performs comprehensive validation before processing to ensure data integrity.

    Args:
        file: Uploaded Excel file object
        user: Current user performing the upload

    Returns:
        dict: Result dictionary with status, message, successes, and errors

    Example Success Response:
        {
            "status": "success",
            "message": "Bulk upload for forms processed successfully.",
            "total_processed": 3,
            "total_success": 3,
            "total_errors": 0,
            "successes": [
                "Form 'Safety Inspection' created with 3 sections and 12 fields.",
                ...
            ],
            "errors": [],
            "created_forms": [
                {
                    "id": "uuid-here",
                    "unique_code": "FORM-00123",
                    "title": "Safety Inspection",
                    "sections_count": 3,
                    "fields_count": 12
                }
            ]
        }

    Error Response Example:
        {
            "status": "failed",
            "message": "Validation failed. Please fix the errors below and re-upload the file.",
            "errors": [
                {
                    "row": 5,
                    "sheet": "Fields",
                    "type": "Invalid Dropdown Value",
                    "column": "Field Type",
                    "invalid_value": "invalid_type",
                    "message": "'invalid_type' is not a valid Field Type."
                },
                ...
            ],
            "total_errors": 5
        }
    """
    try:
        file_name = file.name.lower()

        # Check file format
        if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return {
                "status": "failed",
                "message": "Unsupported file format. Please upload Excel file (.xlsx or .xls)."
            }

        # Parse Excel file (all sheets)
        parsed_data = parse_excel_file_multisheet(file)

        if not parsed_data:
            return {"status": "failed", "message": "No data found in the uploaded file."}

        # Check if we have forms data
        if not parsed_data.get("forms"):
            return {"status": "failed", "message": "No forms found in the Forms sheet."}

        # VALIDATION PHASE - Check all sheets before processing
        print("\n=== Starting Validation ===")
        validation_errors = validate_all_sheets(parsed_data, user)

        if validation_errors:
            print(f"Validation FAILED with {len(validation_errors)} errors")
            for error in validation_errors[:5]:  # Show first 5 errors
                print(f"  - Row {error.get('row')}, Sheet {error.get('sheet')}: {error.get('message')}")
            return {
                "status": "failed",
                "message": "Validation failed. Please fix the errors below and re-upload the file.",
                "errors": validation_errors,
                "total_errors": len(validation_errors)
            }

        print("Validation PASSED! Starting form creation...")

        # PROCESSING PHASE - Only if validation passes
        success_messages = []
        error_messages = []
        created_forms = []

        print(f"\n=== Processing {len(parsed_data['forms'])} forms ===")
        for idx, form_data in enumerate(parsed_data["forms"], start=1):
            print(f"\nProcessing form {idx}/{len(parsed_data['forms'])}")
            print(f"Form data: {form_data}")
            print(f"Calling create_single_form...")
            try:
                result = create_single_form(form_data, parsed_data, user)
                print(f"create_single_form returned: {result.get('status')}")

                if result.get("status") == "success":
                    success_messages.append(result["message"])
                    created_forms.append(result["form_info"])
                else:
                    error_messages.append(result["message"])

            except Exception as e:
                logger.exception(f"Error creating form '{form_data.get('form_title')}': {e}")
                error_messages.append(
                    f"Form '{form_data.get('form_title')}': Unexpected error - {str(e)}"
                )

        return {
            "status": "success" if not error_messages else "partial_success",
            "message": f"Bulk upload for forms processed.",
            "total_processed": len(parsed_data["forms"]),
            "total_success": len(success_messages),
            "total_errors": len(error_messages),
            "successes": success_messages,
            "errors": error_messages,
            "created_forms": created_forms
        }

    except Exception as e:
        logger.exception(f"Error in bulk_upload_forms_service: {e}")
        return {
            "status": "failed",
            "message": f"Error processing file: {str(e)}. Please contact administrator."
        }


# ===============================
# File Parsing Functions
# ===============================

def parse_excel_file_multisheet(file):
    """
    Parse Excel file with multiple sheets and return structured data.

    Expected sheets:
    - Forms: Form metadata
    - Sections: Section definitions
    - Fields: Field definitions
    - References: Dropdown data (ignored during parsing)

    Args:
        file: Excel file object (.xlsx or .xls)

    Returns:
        dict: {
            "forms": [list of form dictionaries],
            "sections": [list of section dictionaries],
            "fields": [list of field dictionaries]
        }

    Raises:
        Exception: If file cannot be read or required sheets are missing
    """
    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        logger.error("Error reading Excel file: %s", e)
        raise Exception("Unable to read Excel file. Please ensure it's a valid Excel file.")

    result = {
        "forms": [],
        "sections": [],
        "fields": []
    }

    # Parse Forms sheet
    if SHEET_NAMES["forms"] in wb.sheetnames:
        result["forms"] = parse_sheet(wb[SHEET_NAMES["forms"]], "forms")
        print(f"\n=== Parsed {len(result['forms'])} forms from Excel ===")
        for form in result["forms"]:
            print(f"Form: {form.get('form_title')}")
    else:
        raise Exception(f"Required sheet '{SHEET_NAMES['forms']}' not found in Excel file.")

    # Parse Sections sheet
    if SHEET_NAMES["sections"] in wb.sheetnames:
        result["sections"] = parse_sheet(wb[SHEET_NAMES["sections"]], "sections")
        print(f"\n=== Parsed {len(result['sections'])} sections from Excel ===")
        for section in result["sections"]:
            print(f"Section: {section.get('section_name')} for form: {section.get('form_title')}")

    # Parse Fields sheet
    if SHEET_NAMES["fields"] in wb.sheetnames:
        result["fields"] = parse_sheet(wb[SHEET_NAMES["fields"]], "fields")
        print(f"\n=== Parsed {len(result['fields'])} fields from Excel ===")
        for field in result["fields"]:
            print(f"Field: {field.get('field_label')} in section: {field.get('section_name')} for form: {field.get('form_title')}")

    return result


def parse_sheet(worksheet, sheet_type):
    """
    Parse a single worksheet and return list of row dictionaries.

    Reads from row 3 (header) and row 4 onwards (data).
    Row 1 contains title, Row 2 contains instructions.

    Args:
        worksheet: openpyxl worksheet object
        sheet_type: Type of sheet ('forms', 'sections', 'fields')

    Returns:
        list: List of dictionaries with normalized column names
    """
    # Read header from row 3
    header = []
    for cell in next(worksheet.iter_rows(min_row=3, max_row=3)):
        header.append(cell.value.strip() if cell.value else "")

    # Normalize column names using COLUMN_MAPPING
    normalized_header = []
    for col in header:
        normalized_header.append(COLUMN_MAPPING.get(col, col.lower().replace(" ", "_")))

    # Read data starting from row 4, skip empty rows
    rows = []
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_dict = dict(zip(normalized_header, row))

        # Add sheet_type for tracking
        row_dict["_sheet_type"] = sheet_type

        rows.append(row_dict)

    return rows


# ===============================
# Validation Functions
# ===============================

def validate_all_sheets(parsed_data, user):
    """
    Validate all sheets for:
    1. Missing required data
    2. Invalid dropdown selections
    3. Cross-sheet reference validation
    4. JSON format validation
    5. Duplicate entries
    6. Field type and data type compatibility

    Args:
        parsed_data (dict): Parsed data from all sheets
        user: Current user

    Returns:
        list: List of validation error dictionaries
    """
    validation_errors = []

    # Fetch valid dropdown values from database
    # FormType uses effective_end_date for soft delete
    valid_form_types = list(
        FormType.objects.filter(effective_end_date__isnull=True)
        .values_list("name", flat=True)
    )

    # FieldType uses is_deleted for soft delete
    valid_field_types = list(
        FieldType.objects.filter(is_deleted=False)
        .values_list("name", flat=True)
    )

    # DataType uses is_deleted for soft delete
    valid_data_types = list(
        DataType.objects.filter(is_deleted=False)
        .values_list("name", flat=True)
    )

    # Get field type to data type mapping
    field_type_data_type_map = {}
    for ft in FieldType.objects.filter(is_deleted=False).select_related('data_type'):
        field_type_data_type_map[ft.name] = ft.data_type.name

    # Track seen values for duplicate detection
    seen_form_titles = set()
    seen_sections = set()  # (form_title, section_name)
    seen_section_orders = set()  # (form_title, section_order)
    seen_fields = set()  # (form_title, section_name, field_label)
    seen_field_orders = set()  # (form_title, section_name, field_order)

    # Build form titles from Forms sheet for cross-sheet validation
    form_titles_in_file = {str(row.get("form_title")).strip() for row in parsed_data.get("forms", [])
                           if row.get("form_title") and str(row.get("form_title")).strip()}

    # Build sections map for cross-sheet validation: {form_title: [section_names]}
    sections_map = {}
    for row in parsed_data.get("sections", []):
        form_title = row.get("form_title")
        section_name = row.get("section_name")
        if form_title and section_name:
            form_title_str = str(form_title).strip()
            section_name_str = str(section_name).strip()
            if form_title_str and section_name_str:
                if form_title_str not in sections_map:
                    sections_map[form_title_str] = set()
                sections_map[form_title_str].add(section_name_str)

    # =========================
    # VALIDATE FORMS SHEET
    # =========================
    for idx, row_data in enumerate(parsed_data.get("forms", []), start=1):
        actual_row = idx + 3  # Row 1=Title, Row 2=Note, Row 3=Header, Data starts at 4

        # 1. Check required fields
        missing_fields = []
        for field in REQUIRED_FIELDS["forms"]:
            value = row_data.get(field)
            if value is None or str(value).strip() == "":
                original_col_name = get_original_column_name(field)
                missing_fields.append(original_col_name)

        if missing_fields:
            validation_errors.append({
                "row": actual_row,
                "sheet": SHEET_NAMES["forms"],
                "type": "Missing Data",
                "message": f"Missing required data in column(s): {', '.join(missing_fields)}"
            })

        # 2. Validate dropdown: form_type
        form_type = row_data.get("form_type")
        if form_type and str(form_type).strip():
            form_type_str = str(form_type).strip()
            if not any(form_type_str.lower() == ft.lower() for ft in valid_form_types):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["forms"],
                    "type": "Invalid Dropdown Value",
                    "column": "Form Type",
                    "invalid_value": form_type_str,
                    "message": f"'{form_type_str}' is not a valid Form Type. Must be selected from dropdown."
                })

        # 3. Check duplicate form_title in file
        form_title = row_data.get("form_title")
        if form_title and str(form_title).strip():
            form_title_str = str(form_title).strip()

            if form_title_str.lower() in seen_form_titles:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["forms"],
                    "type": "Duplicate in File",
                    "column": "Form Title",
                    "invalid_value": form_title_str,
                    "message": f"Form title '{form_title_str}' appears multiple times in this file."
                })
            else:
                seen_form_titles.add(form_title_str.lower())

            # 4. Check duplicate form_title in database
            if Form.objects.filter(
                title__iexact=form_title_str,
                effective_end_date__isnull=True
            ).exists():
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["forms"],
                    "type": "Duplicate Entry",
                    "column": "Form Title",
                    "invalid_value": form_title_str,
                    "message": f"Form with title '{form_title_str}' already exists in database."
                })

        # 5. Validate is_completed field (should be TRUE/FALSE or boolean)
        is_completed = row_data.get("is_completed")
        if is_completed is not None and str(is_completed).strip():
            is_completed_str = str(is_completed).strip().upper()
            if is_completed_str not in ["TRUE", "FALSE", "YES", "NO", "1", "0"]:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["forms"],
                    "type": "Invalid Format",
                    "column": "Is Completed",
                    "invalid_value": str(is_completed),
                    "message": f"'{is_completed}' is not a valid boolean. Use TRUE or FALSE."
                })

    # =========================
    # VALIDATE SECTIONS SHEET
    # =========================
    for idx, row_data in enumerate(parsed_data.get("sections", []), start=1):
        actual_row = idx + 3

        # 1. Check required fields
        missing_fields = []
        for field in REQUIRED_FIELDS["sections"]:
            value = row_data.get(field)
            if value is None or str(value).strip() == "":
                original_col_name = get_original_column_name(field)
                missing_fields.append(original_col_name)

        if missing_fields:
            validation_errors.append({
                "row": actual_row,
                "sheet": SHEET_NAMES["sections"],
                "type": "Missing Data",
                "message": f"Missing required data in column(s): {', '.join(missing_fields)}"
            })

        # 2. Validate form_title exists in Forms sheet
        form_title = row_data.get("form_title")
        if form_title and str(form_title).strip():
            form_title_str = str(form_title).strip()
            if form_title_str not in form_titles_in_file:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["sections"],
                    "type": "Invalid Reference",
                    "column": "Form Title",
                    "invalid_value": form_title_str,
                    "message": f"Form '{form_title_str}' not found in Forms sheet."
                })

        # 3. Validate section_order is a positive integer
        section_order = row_data.get("section_order")
        if section_order is not None:
            try:
                order_int = int(section_order)
                if order_int <= 0:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["sections"],
                        "type": "Invalid Format",
                        "column": "Section Order",
                        "invalid_value": str(section_order),
                        "message": "Section Order must be a positive integer."
                    })
            except (ValueError, TypeError):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["sections"],
                    "type": "Invalid Format",
                    "column": "Section Order",
                    "invalid_value": str(section_order),
                    "message": "Section Order must be a valid integer."
                })

        # 4. Check duplicate (form_title, section_name)
        section_name = row_data.get("section_name")
        if form_title and section_name:
            form_title_str = str(form_title).strip()
            section_name_str = str(section_name).strip()
            section_key = (form_title_str.lower(), section_name_str.lower())

            if section_key in seen_sections:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["sections"],
                    "type": "Duplicate in File",
                    "column": "Section Name",
                    "invalid_value": section_name_str,
                    "message": f"Section '{section_name_str}' appears multiple times for form '{form_title_str}'."
                })
            else:
                seen_sections.add(section_key)

        # 5. Check duplicate (form_title, section_order)
        if form_title and section_order is not None:
            form_title_str = str(form_title).strip()
            try:
                order_int = int(section_order)
                order_key = (form_title_str.lower(), order_int)

                if order_key in seen_section_orders:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["sections"],
                        "type": "Duplicate in File",
                        "column": "Section Order",
                        "invalid_value": str(section_order),
                        "message": f"Section Order {section_order} is used multiple times for form '{form_title_str}'."
                    })
                else:
                    seen_section_orders.add(order_key)
            except (ValueError, TypeError):
                pass  # Already handled above

        # 6. Validate simple dependency columns
        dep_section = row_data.get("dependency_section")
        dep_field = row_data.get("dependency_field")
        dep_option = row_data.get("dependency_option")

        # Check if any of the simple dependency columns are filled
        has_dep_section = dep_section and str(dep_section).strip()
        has_dep_field = dep_field and str(dep_field).strip()
        has_dep_option = dep_option and str(dep_option).strip()

        # If any one is provided, all three must be provided
        if has_dep_section or has_dep_field or has_dep_option:
            missing_dep_cols = []
            if not has_dep_section:
                missing_dep_cols.append("Dependency Section")
            if not has_dep_field:
                missing_dep_cols.append("Dependency Field")
            if not has_dep_option:
                missing_dep_cols.append("Dependency Option")

            if missing_dep_cols:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["sections"],
                    "type": "Missing Data",
                    "message": f"When using simple dependency, all three columns must be filled: {', '.join(missing_dep_cols)}"
                })
            else:
                # All three are provided - validate them
                dep_section_str = str(dep_section).strip()
                dep_field_str = str(dep_field).strip()
                dep_option_str = str(dep_option).strip()

                # Validate dependency section exists in the same form
                if form_title:
                    form_title_str = str(form_title).strip()
                    if form_title_str in sections_map:
                        if dep_section_str not in sections_map[form_title_str]:
                            validation_errors.append({
                                "row": actual_row,
                                "sheet": SHEET_NAMES["sections"],
                                "type": "Invalid Reference",
                                "column": "Dependency Section",
                                "invalid_value": dep_section_str,
                                "message": f"Dependency section '{dep_section_str}' not found in Sections sheet for form '{form_title_str}'."
                            })

                # Validate dependency field exists in the dependency section
                field_found = False
                field_options = []
                for field_row in parsed_data.get("fields", []):
                    if (str(field_row.get("form_title", "")).strip() == form_title_str and
                        str(field_row.get("section_name", "")).strip() == dep_section_str and
                        str(field_row.get("field_label", "")).strip() == dep_field_str):
                        field_found = True
                        # Get field options if available
                        options_str = field_row.get("options")
                        if options_str and str(options_str).strip():
                            try:
                                field_options = json.loads(str(options_str).strip())
                            except:
                                pass
                        break

                if not field_found:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["sections"],
                        "type": "Invalid Reference",
                        "column": "Dependency Field",
                        "invalid_value": dep_field_str,
                        "message": f"Dependency field '{dep_field_str}' not found in section '{dep_section_str}' in Fields sheet."
                    })
                elif field_options and dep_option_str not in field_options:
                    # Validate option exists in field's options (if options are defined)
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["sections"],
                        "type": "Invalid Reference",
                        "column": "Dependency Option",
                        "invalid_value": dep_option_str,
                        "message": f"Dependency option '{dep_option_str}' not found in field '{dep_field_str}' options: {field_options}"
                    })

        # 7. Validate dependency JSON format (only if simple dependency not used)
        dependency = row_data.get("dependency")
        if dependency and str(dependency).strip() and not (has_dep_section and has_dep_field and has_dep_option):
            dependency_str = str(dependency).strip()
            if not is_valid_json(dependency_str):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["sections"],
                    "type": "Invalid Format",
                    "column": "Dependency (JSON)",
                    "invalid_value": dependency_str,
                    "message": "Invalid JSON format for Dependency field."
                })

    # =========================
    # VALIDATE FIELDS SHEET
    # =========================
    for idx, row_data in enumerate(parsed_data.get("fields", []), start=1):
        actual_row = idx + 3

        # 1. Check required fields
        missing_fields = []
        for field in REQUIRED_FIELDS["fields"]:
            value = row_data.get(field)
            if value is None or str(value).strip() == "":
                original_col_name = get_original_column_name(field)
                missing_fields.append(original_col_name)

        if missing_fields:
            validation_errors.append({
                "row": actual_row,
                "sheet": SHEET_NAMES["fields"],
                "type": "Missing Data",
                "message": f"Missing required data in column(s): {', '.join(missing_fields)}"
            })

        # 2. Validate form_title exists in Forms sheet
        form_title = row_data.get("form_title")
        if form_title and str(form_title).strip():
            form_title_str = str(form_title).strip()
            if form_title_str not in form_titles_in_file:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Reference",
                    "column": "Form Title",
                    "invalid_value": form_title_str,
                    "message": f"Form '{form_title_str}' not found in Forms sheet."
                })

        # 3. Validate section_name exists in Sections sheet for this form
        section_name = row_data.get("section_name")
        if form_title and section_name:
            form_title_str = str(form_title).strip()
            section_name_str = str(section_name).strip()

            if form_title_str in sections_map:
                if section_name_str not in sections_map[form_title_str]:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["fields"],
                        "type": "Invalid Reference",
                        "column": "Section Name",
                        "invalid_value": section_name_str,
                        "message": f"Section '{section_name_str}' not found for form '{form_title_str}' in Sections sheet."
                    })

        # 4. Validate dropdowns: field_type and data_type
        field_type = row_data.get("field_type")
        if field_type and str(field_type).strip():
            field_type_str = str(field_type).strip()
            if not any(field_type_str.lower() == ft.lower() for ft in valid_field_types):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Dropdown Value",
                    "column": "Field Type",
                    "invalid_value": field_type_str,
                    "message": f"'{field_type_str}' is not a valid Field Type. Must be selected from dropdown."
                })

        data_type = row_data.get("data_type")
        if data_type and str(data_type).strip():
            data_type_str = str(data_type).strip()
            if not any(data_type_str.lower() == dt.lower() for dt in valid_data_types):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Dropdown Value",
                    "column": "Data Type",
                    "invalid_value": data_type_str,
                    "message": f"'{data_type_str}' is not a valid Data Type. Must be selected from dropdown."
                })

        # 5. Validate field_type and data_type compatibility
        if field_type and data_type:
            field_type_str = str(field_type).strip()
            data_type_str = str(data_type).strip()

            expected_data_type = field_type_data_type_map.get(field_type_str)
            if expected_data_type and expected_data_type.lower() != data_type_str.lower():
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Incompatible Types",
                    "column": "Data Type",
                    "invalid_value": data_type_str,
                    "message": f"Field Type '{field_type_str}' requires Data Type '{expected_data_type}', but '{data_type_str}' was provided."
                })

        # 6. Validate field_order is a positive integer
        field_order = row_data.get("field_order")
        if field_order is not None:
            try:
                order_int = int(field_order)
                if order_int <= 0:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["fields"],
                        "type": "Invalid Format",
                        "column": "Field Order",
                        "invalid_value": str(field_order),
                        "message": "Field Order must be a positive integer."
                    })
            except (ValueError, TypeError):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Format",
                    "column": "Field Order",
                    "invalid_value": str(field_order),
                    "message": "Field Order must be a valid integer."
                })

        # 7. Check duplicate (form_title, section_name, field_label)
        field_label = row_data.get("field_label")
        if form_title and section_name and field_label:
            form_title_str = str(form_title).strip()
            section_name_str = str(section_name).strip()
            field_label_str = str(field_label).strip()
            field_key = (form_title_str.lower(), section_name_str.lower(), field_label_str.lower())

            if field_key in seen_fields:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Duplicate in File",
                    "column": "Field Label",
                    "invalid_value": field_label_str,
                    "message": f"Field '{field_label_str}' appears multiple times in section '{section_name_str}' of form '{form_title_str}'."
                })
            else:
                seen_fields.add(field_key)

        # 8. Check duplicate (form_title, section_name, field_order)
        if form_title and section_name and field_order is not None:
            form_title_str = str(form_title).strip()
            section_name_str = str(section_name).strip()
            try:
                order_int = int(field_order)
                order_key = (form_title_str.lower(), section_name_str.lower(), order_int)

                if order_key in seen_field_orders:
                    validation_errors.append({
                        "row": actual_row,
                        "sheet": SHEET_NAMES["fields"],
                        "type": "Duplicate in File",
                        "column": "Field Order",
                        "invalid_value": str(field_order),
                        "message": f"Field Order {field_order} is used multiple times in section '{section_name_str}' of form '{form_title_str}'."
                    })
                else:
                    seen_field_orders.add(order_key)
            except (ValueError, TypeError):
                pass  # Already handled above

        # 9. Validate options JSON format (for dropdowns, checkboxes, etc.)
        options = row_data.get("options")
        if options and str(options).strip():
            options_str = str(options).strip()
            if not is_valid_json(options_str):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Format",
                    "column": "Options",
                    "invalid_value": options_str,
                    "message": "Invalid JSON format. Expected array like: [\"Option1\",\"Option2\"]"
                })

        # 10. Validate simple field dependency columns
        field_dep_section = row_data.get("field_dep_section")
        field_dep_field = row_data.get("field_dep_field")
        field_dep_option = row_data.get("field_dep_option")

        # Check if any of the simple field dependency columns are filled
        has_field_dep_section = field_dep_section and str(field_dep_section).strip()
        has_field_dep_field = field_dep_field and str(field_dep_field).strip()
        has_field_dep_option = field_dep_option and str(field_dep_option).strip()

        # If any one is provided, all three must be provided
        if has_field_dep_section or has_field_dep_field or has_field_dep_option:
            missing_field_dep_cols = []
            if not has_field_dep_section:
                missing_field_dep_cols.append("Field Dep Section")
            if not has_field_dep_field:
                missing_field_dep_cols.append("Field Dep Field")
            if not has_field_dep_option:
                missing_field_dep_cols.append("Field Dep Option")

            if missing_field_dep_cols:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Missing Data",
                    "message": f"When using simple field dependency, all three columns must be filled: {', '.join(missing_field_dep_cols)}"
                })
            else:
                # All three are provided - validate them
                field_dep_section_str = str(field_dep_section).strip()
                field_dep_field_str = str(field_dep_field).strip()
                field_dep_option_str = str(field_dep_option).strip()

                # Validate dependency section exists in the same form
                if form_title:
                    form_title_str = str(form_title).strip()
                    if form_title_str in sections_map:
                        if field_dep_section_str not in sections_map[form_title_str]:
                            validation_errors.append({
                                "row": actual_row,
                                "sheet": SHEET_NAMES["fields"],
                                "type": "Invalid Reference",
                                "column": "Field Dep Section",
                                "invalid_value": field_dep_section_str,
                                "message": f"Field dependency section '{field_dep_section_str}' not found in Sections sheet for form '{form_title_str}'."
                            })

                    # Validate dependency field exists in the dependency section
                    dep_field_found = False
                    dep_field_options = []
                    for field_row in parsed_data.get("fields", []):
                        if (str(field_row.get("form_title", "")).strip() == form_title_str and
                            str(field_row.get("section_name", "")).strip() == field_dep_section_str and
                            str(field_row.get("field_label", "")).strip() == field_dep_field_str):
                            dep_field_found = True
                            # Get field options if available
                            options_str = field_row.get("options")
                            if options_str and str(options_str).strip():
                                try:
                                    dep_field_options = json.loads(str(options_str).strip())
                                except:
                                    pass
                            break

                    if not dep_field_found:
                        validation_errors.append({
                            "row": actual_row,
                            "sheet": SHEET_NAMES["fields"],
                            "type": "Invalid Reference",
                            "column": "Field Dep Field",
                            "invalid_value": field_dep_field_str,
                            "message": f"Field dependency field '{field_dep_field_str}' not found in section '{field_dep_section_str}' in Fields sheet."
                        })
                    elif dep_field_options and field_dep_option_str not in dep_field_options:
                        # Validate option exists in field's options (if options are defined)
                        validation_errors.append({
                            "row": actual_row,
                            "sheet": SHEET_NAMES["fields"],
                            "type": "Invalid Reference",
                            "column": "Field Dep Option",
                            "invalid_value": field_dep_option_str,
                            "message": f"Field dependency option '{field_dep_option_str}' not found in field '{field_dep_field_str}' options: {dep_field_options}"
                        })

        # 11. Validate dependency JSON format (only if simple columns not used)
        dependency = row_data.get("field_dependency")
        if dependency and str(dependency).strip() and not (has_field_dep_section and has_field_dep_field and has_field_dep_option):
            dependency_str = str(dependency).strip()
            if not is_valid_json(dependency_str):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Format",
                    "column": "Dependency",
                    "invalid_value": dependency_str,
                    "message": "Invalid JSON format for Dependency field."
                })

        # 11. Validate additional_info JSON format
        additional_info = row_data.get("additional_info")
        if additional_info and str(additional_info).strip():
            additional_info_str = str(additional_info).strip()
            if not is_valid_json(additional_info_str):
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Format",
                    "column": "Additional Info",
                    "invalid_value": additional_info_str,
                    "message": "Invalid JSON format for Additional Info field."
                })

        # 12. Validate required field (should be TRUE/FALSE or boolean)
        required = row_data.get("required")
        if required is not None and str(required).strip():
            required_str = str(required).strip().upper()
            if required_str not in ["TRUE", "FALSE", "YES", "NO", "1", "0"]:
                validation_errors.append({
                    "row": actual_row,
                    "sheet": SHEET_NAMES["fields"],
                    "type": "Invalid Format",
                    "column": "Required",
                    "invalid_value": str(required),
                    "message": f"'{required}' is not a valid boolean. Use TRUE or FALSE."
                })

    return validation_errors


def get_original_column_name(normalized_field):
    """
    Get the original column name from normalized field name.

    Args:
        normalized_field (str): Normalized field name (e.g., 'form_title')

    Returns:
        str: Original column name (e.g., 'Form Title')
    """
    for original, normalized in COLUMN_MAPPING.items():
        if normalized == normalized_field:
            return original
    return normalized_field.replace("_", " ").title()


def is_valid_json(json_string):
    """
    Validate if a string is valid JSON.

    Args:
        json_string (str): String to validate

    Returns:
        bool: True if valid JSON, False otherwise
    """
    try:
        json.loads(json_string)
        return True
    except (json.JSONDecodeError, TypeError):
        return False


def parse_boolean(value):
    """
    Parse various boolean representations to Python boolean.

    Args:
        value: Value to parse (can be string, int, bool)

    Returns:
        bool: Parsed boolean value (defaults to False if invalid)
    """
    if value is None:
        return False

    if isinstance(value, bool):
        return value

    value_str = str(value).strip().upper()
    return value_str in ["TRUE", "YES", "1"]


def generate_unique_field_name(prefix="field"):
    """
    Generate a unique field name matching the frontend logic.
    Uses timestamp and random string to ensure uniqueness.

    Frontend equivalent:
    const generateUniqueName = (prefix = "field") => {
      const timestamp = Date.now();
      const random = Math.random().toString(36).substring(2, 8);
      return `${prefix}_${timestamp}_${random}`;
    };

    Args:
        prefix (str): Prefix for the field name (default: "field")

    Returns:
        str: Unique field name in format: {prefix}_{timestamp}_{random}
    """
    import time
    import random
    import string

    # Get current timestamp in milliseconds (like Date.now() in JavaScript)
    timestamp = int(time.time() * 1000)

    # Generate random 6-character string (like Math.random().toString(36).substring(2, 8))
    # Using lowercase letters and digits to match JS toString(36)
    random_str = ''.join(random.choices(string.ascii_lowercase + string.digits, k=6))

    return f"{prefix}_{timestamp}_{random_str}"


# ===============================
# Form Creation Functions
# ===============================

@transaction.atomic
def create_single_form(form_data, parsed_data, user):
    """
    Create a single form with its sections and fields in an atomic transaction.

    Args:
        form_data (dict): Form metadata from Forms sheet
        parsed_data (dict): All parsed data (forms, sections, fields)
        user: Current user

    Returns:
        dict: Result with status, message, and form_info
    """
    try:
        print(f"\n>>> Entering create_single_form function")

        form_title = str(form_data.get("form_title", "")).strip()
        form_type_name = str(form_data.get("form_type", "")).strip()
        description = str(form_data.get("description", "")).strip() if form_data.get("description") else ""
        is_completed = parse_boolean(form_data.get("is_completed", False))

        print(f">>> Form title: {form_title}")
        print(f">>> Form type: {form_type_name}")
        print(f">>> Looking up FormType...")

        # Get FormType
        form_type = FormType.objects.filter(
            name__iexact=form_type_name,
            effective_end_date__isnull=True
        ).first()

        if not form_type:
            print(f">>> ERROR: FormType '{form_type_name}' not found")
            return {
                "status": "failed",
                "message": f"Form '{form_title}': Form Type '{form_type_name}' not found."
            }

        print(f">>> FormType found: {form_type.name} (ID: {form_type.id})")
        print(f">>> Creating Form object...")

        # Create Form
        form = Form.objects.create(
            title=form_title,
            form_type=form_type,
            description=description,
            is_completed=is_completed,
            created_by=user,
            root_form=None  # Will be set to itself after creation
        )

        print(f">>> Form created: {form.id}")

        # Set root_form to itself (first version)
        form.root_form = form
        form.save()
        print(f">>> root_form set to itself")

        # Get sections for this form
        form_sections = [
            s for s in parsed_data.get("sections", [])
            if str(s.get("form_title", "")).strip().lower() == form_title.lower()
        ]

        # Get fields for this form
        form_fields = [
            f for f in parsed_data.get("fields", [])
            if str(f.get("form_title", "")).strip().lower() == form_title.lower()
        ]

        # Debug output
        print(f"\n=== Creating form '{form_title}' ===")
        print(f"Found {len(form_sections)} sections")
        print(f"Found {len(form_fields)} fields")
        print(f"All sections in parsed_data: {len(parsed_data.get('sections', []))}")
        print(f"All fields in parsed_data: {len(parsed_data.get('fields', []))}")

        # Create sections
        section_map = {}  # Maps section_name to FormSections object
        sections_created = 0

        for section_data in form_sections:
            section_name = str(section_data.get("section_name", "")).strip()
            section_description = str(section_data.get("section_description", "")).strip() if section_data.get("section_description") else ""
            section_order = int(section_data.get("section_order", 1))

            # Get simple dependency columns
            dep_section = str(section_data.get("dependency_section", "")).strip() if section_data.get("dependency_section") else None
            dep_field = str(section_data.get("dependency_field", "")).strip() if section_data.get("dependency_field") else None
            dep_option = str(section_data.get("dependency_option", "")).strip() if section_data.get("dependency_option") else None

            # Get JSON dependency column
            dependency_str = str(section_data.get("dependency", "")).strip() if section_data.get("dependency") else None

            print(f"Creating section: {section_name}, order: {section_order}")

            dependency = None

            # Priority 1: Use simple dependency columns if all three are provided
            if dep_section and dep_field and dep_option:
                print(f"Building dependency from simple columns: Section={dep_section}, Field={dep_field}, Option={dep_option}")

                # Find the field in the parsed data to get its field_name
                field_name_to_use = None
                for field in form_fields:
                    if (str(field.get("section_name", "")).strip().lower() == dep_section.lower() and
                        str(field.get("field_label", "")).strip().lower() == dep_field.lower()):
                        # Use the actual field_name from Excel (not the label)
                        field_name_to_use = field.get("field_name", "")
                        break

                if not field_name_to_use:
                    # If field not found in current form, use the provided field label as fallback
                    # This will be corrected later in the update phase
                    field_name_to_use = dep_field

                # Build the dependency JSON structure
                dependency = {
                    "field_name": field_name_to_use,  # Will be updated after field creation
                    "field_section": dep_section,
                    "options_selected": [dep_option],
                    "cascader_selection": [[dep_section, field_name_to_use, dep_option]],
                    "multiple_field_dependencies": [
                        {
                            "field_name": field_name_to_use,
                            "field_section": dep_section,
                            "options_selected": [dep_option]
                        }
                    ]
                }
                print(f"Generated dependency JSON: {dependency}")

            # Priority 2: Use JSON dependency if provided and simple columns not used
            elif dependency_str:
                try:
                    dependency = json.loads(dependency_str)
                except json.JSONDecodeError:
                    dependency = None

            section = FormSections.objects.create(
                form=form,
                name=section_name,
                description=section_description,
                order=section_order,
                dependency=dependency
            )

            print(f"Section created successfully: {section.id}")

            section_map[section_name.lower()] = section
            sections_created += 1

        # Create fields
        fields_created = 0
        parent_field_map = {}  # Maps (section, field_label) to FormFields object

        # Sort fields to ensure parent fields are created before children
        sorted_fields = sorted(
            form_fields,
            key=lambda f: (1 if f.get("parent_field") else 0)  # Non-parent fields first
        )

        for field_data in sorted_fields:
            section_name = str(field_data.get("section_name", "")).strip()
            field_label = str(field_data.get("field_label", "")).strip()
            field_type_name = str(field_data.get("field_type", "")).strip()
            required = parse_boolean(field_data.get("required", False))
            field_order = int(field_data.get("field_order", 1))

            # Auto-generate unique field name (matching frontend logic)
            field_name = generate_unique_field_name()

            print(f"Creating field: {field_label} (name: {field_name}) in section: {section_name}")

            # Get section
            section = section_map.get(section_name.lower())
            if not section:
                print(f"ERROR: Section '{section_name}' not found in section_map")
                print(f"Available sections: {list(section_map.keys())}")
                continue  # Skip if section not found (should be caught in validation)

            # Get FieldType (uses is_deleted for soft delete)
            field_type = FieldType.objects.filter(
                name__iexact=field_type_name,
                is_deleted=False
            ).first()

            if not field_type:
                print(f"ERROR: FieldType '{field_type_name}' not found")
                continue  # Skip if field type not found

            # Parse additional_info
            additional_info_str = str(field_data.get("additional_info", "")).strip() if field_data.get("additional_info") else None
            additional_info = {}

            if additional_info_str:
                try:
                    additional_info = json.loads(additional_info_str)
                except json.JSONDecodeError:
                    additional_info = {}

            # Parse options and add to additional_info
            options_str = str(field_data.get("options", "")).strip() if field_data.get("options") else None
            if options_str:
                try:
                    options = json.loads(options_str)
                    additional_info["options"] = options
                except json.JSONDecodeError:
                    pass

            # Parse width and add to additional_info (default to "100" if not provided)
            width_str = str(field_data.get("width", "")).strip() if field_data.get("width") else None
            if width_str:
                # Extract numeric part from format like "25% (1/4)" or "100% (Full)"
                # Split by '%' and take the first part
                width_value = width_str.split('%')[0].strip()
                additional_info["width"] = width_value
            else:
                additional_info["width"] = "100"  # Default to 100% width

            # Parse validation rules and add to additional_info
            validation_str = str(field_data.get("validation", "")).strip() if field_data.get("validation") else None
            if validation_str:
                try:
                    validation = json.loads(validation_str)
                    additional_info["validation"] = validation
                    print(f"Added validation rules for field '{field_label}': {validation}")
                except json.JSONDecodeError:
                    print(f"Warning: Invalid JSON in Validation column for field '{field_label}': {validation_str}")

            # Auto-configure dynamic fields based on field type name or label
            field_type_lower = field_type.name.lower()
            field_label_lower = field_label.lower()

            # Define common dynamic field mappings
            dynamic_mappings = {
                "location": "/api/config/locations/",
                "department": "/api/config/departments/",
                "designation": "/api/config/designations/",
                "role": "/api/config/roles/",
                "employee": "/api/config/employees/list/",
                "user": "/api/config/employees/list/",
            }

            # Check if this field should be dynamic based on its name or label
            for keyword, api_endpoint in dynamic_mappings.items():
                if keyword in field_type_lower or keyword in field_label_lower:
                    # Store dynamic configuration in additional_info
                    additional_info["dynamic"] = True
                    additional_info["end_point"] = api_endpoint
                    print(f"Auto-configured field '{field_label}' as dynamic with endpoint: {api_endpoint}")
                    break

            # Parse field dependency - Priority 1: Simple columns, Priority 2: JSON
            field_dep_section = str(field_data.get("field_dep_section", "")).strip() if field_data.get("field_dep_section") else None
            field_dep_field = str(field_data.get("field_dep_field", "")).strip() if field_data.get("field_dep_field") else None
            field_dep_option = str(field_data.get("field_dep_option", "")).strip() if field_data.get("field_dep_option") else None

            dependency = None

            # Priority 1: Use simple field dependency columns if all three are provided
            if field_dep_section and field_dep_field and field_dep_option:
                print(f"Building field dependency from simple columns: Section={field_dep_section}, Field={field_dep_field}, Option={field_dep_option}")

                # Build the dependency JSON structure (field_name will be updated after all fields are created)
                dependency = {
                    "field_name": field_dep_field,  # Will be updated with actual field name after creation
                    "field_section": field_dep_section,
                    "options_selected": [field_dep_option],
                    "cascader_selection": [[field_dep_section, field_dep_field, field_dep_option]],
                    "multiple_field_dependencies": [
                        {
                            "field_name": field_dep_field,
                            "field_section": field_dep_section,
                            "options_selected": [field_dep_option]
                        }
                    ]
                }
                print(f"Generated field dependency JSON: {dependency}")

            # Priority 2: Use JSON dependency if provided and simple columns not used
            else:
                dependency_str = str(field_data.get("field_dependency", "")).strip() if field_data.get("field_dependency") else None
                if dependency_str:
                    try:
                        dependency = json.loads(dependency_str)
                    except json.JSONDecodeError:
                        dependency = None

            # Get parent field if specified
            parent_field = None
            parent_field_label = str(field_data.get("parent_field", "")).strip() if field_data.get("parent_field") else None
            if parent_field_label:
                parent_key = (section_name.lower(), parent_field_label.lower())
                parent_field = parent_field_map.get(parent_key)

            # Create field
            field = FormFields.objects.create(
                label=field_label,
                name=field_name,
                field_type=field_type,
                section=section,
                required=required,
                order=field_order,
                additional_info=additional_info,
                parent_field=parent_field,
                dependency=dependency
            )

            print(f"Field created successfully: {field.id}")

            # Store in map for parent field lookup
            field_key = (section_name.lower(), field_label.lower())
            parent_field_map[field_key] = field
            fields_created += 1

        # ==========================================
        # UPDATE SECTION DEPENDENCIES WITH ACTUAL FIELD NAMES
        # ==========================================
        # Now that all fields are created, update section dependencies to use actual field names
        # instead of the temporary field labels
        print("\n>>> Updating section dependencies with actual field names...")
        for section in FormSections.objects.filter(form=form, is_deleted=False):
            if section.dependency and section.dependency.get("field_name"):
                dependency_updated = False
                dep = section.dependency.copy()

                # Get the field identifier from dependency (could be name or label)
                field_identifier = dep.get("field_name", "")
                dep_section_name = dep.get("field_section", "")

                if field_identifier and dep_section_name:
                    # Try to find the field by label first (most common case for bulk upload)
                    field_key = (dep_section_name.lower(), field_identifier.lower())
                    actual_field = parent_field_map.get(field_key)

                    # If not found by label, try to find by actual field name
                    if not actual_field:
                        dep_section = FormSections.objects.filter(
                            form=form,
                            name=dep_section_name,
                            is_deleted=False
                        ).first()

                        if dep_section:
                            actual_field = FormFields.objects.filter(
                                section=dep_section,
                                name=field_identifier,
                                is_deleted=False
                            ).first()

                    if actual_field and actual_field.name != field_identifier:
                        # Update field_name to use the actual field's NAME (from FormFields.name)
                        # This is the name attribute, not the label
                        dep["field_name"] = actual_field.name

                        # Update cascader_selection
                        if dep.get("cascader_selection"):
                            for cascader in dep["cascader_selection"]:
                                if len(cascader) >= 2 and cascader[1] == field_identifier:
                                    cascader[1] = actual_field.name

                        # Update multiple_field_dependencies
                        if dep.get("multiple_field_dependencies"):
                            for multi_dep in dep["multiple_field_dependencies"]:
                                if multi_dep.get("field_name") == field_identifier:
                                    multi_dep["field_name"] = actual_field.name

                        section.dependency = dep
                        section.save()
                        dependency_updated = True
                        print(f">>> Updated section '{section.name}' dependency to use field name: {actual_field.name}")

                if not dependency_updated:
                    print(f">>> Section '{section.name}' dependency not updated (field not found or no dependency)")

        # ==========================================
        # UPDATE FIELD DEPENDENCIES WITH ACTUAL FIELD NAMES
        # ==========================================
        # Now that all fields are created, update field dependencies to use actual field names
        # instead of the temporary field labels
        print("\n>>> Updating field dependencies with actual field names...")
        for section in FormSections.objects.filter(form=form, is_deleted=False):
            for field in FormFields.objects.filter(section=section, is_deleted=False):
                if field.dependency and field.dependency.get("field_name"):
                    field_dep_updated = False
                    dep = field.dependency.copy()

                    # Get the field identifier from dependency (could be name or label)
                    field_identifier = dep.get("field_name", "")
                    dep_section_name = dep.get("field_section", "")

                    if field_identifier and dep_section_name:
                        # Try to find the dependency field by label first (most common case for bulk upload)
                        field_key = (dep_section_name.lower(), field_identifier.lower())
                        actual_dep_field = parent_field_map.get(field_key)

                        # If not found by label, try to find by actual field name
                        if not actual_dep_field:
                            dep_section = FormSections.objects.filter(
                                form=form,
                                name=dep_section_name,
                                is_deleted=False
                            ).first()

                            if dep_section:
                                actual_dep_field = FormFields.objects.filter(
                                    section=dep_section,
                                    name=field_identifier,
                                    is_deleted=False
                                ).first()

                        if actual_dep_field and actual_dep_field.name != field_identifier:
                            # Update field_name to use the actual field's NAME (from FormFields.name)
                            dep["field_name"] = actual_dep_field.name

                            # Update cascader_selection
                            if dep.get("cascader_selection"):
                                for cascader in dep["cascader_selection"]:
                                    if len(cascader) >= 2 and cascader[1] == field_identifier:
                                        cascader[1] = actual_dep_field.name

                            # Update multiple_field_dependencies
                            if dep.get("multiple_field_dependencies"):
                                for multi_dep in dep["multiple_field_dependencies"]:
                                    if multi_dep.get("field_name") == field_identifier:
                                        multi_dep["field_name"] = actual_dep_field.name

                            field.dependency = dep
                            field.save()
                            field_dep_updated = True
                            print(f">>> Updated field '{field.label}' dependency to use field name: {actual_dep_field.name}")

                    if not field_dep_updated and field.dependency:
                        print(f">>> Field '{field.label}' dependency not updated (dependency field not found)")

        # Build draft_data structure matching frontend format
        draft_data = {
            "fields": [],  # Root level fields (empty for now)
            "sections": [],
            "version_id": str(form.id),
            "form_details": {
                "title": form.title,
                "form_type": form_type.unique_code,
                "description": form.description or ""
            }
        }

        # Get all sections with their fields from database
        for section in FormSections.objects.filter(form=form).order_by('order'):
            # Generate section_id using timestamp-based unique identifier
            import time
            section_id = f"section_{int(time.time() * 1000)}_{str(section.id)[:8]}"

            section_data = {
                "section_id": section_id,
                "section_name": section.name,
                "dependency": section.dependency if section.dependency else {
                    "field_name": "",
                    "field_section": "",
                    "options_selected": [],
                    "cascader_selection": [],
                    "multiple_field_dependencies": []
                },
                "fields": []
            }

            # Get all fields for this section
            for field in FormFields.objects.filter(section=section).order_by('order'):
                # Generate field name using timestamp-based unique identifier
                field_name = f"field_{int(time.time() * 1000)}_{str(field.id)[:8]}"

                # Determine if field is dynamic and get endpoint
                is_dynamic = field.field_type.dynamic if field.field_type else False
                endpoint = field.field_type.endpoint if (field.field_type and field.field_type.dynamic) else None

                # Auto-configure dynamic fields based on field type name or label
                field_type_lower = field.field_type.name.lower()
                field_label_lower = field.label.lower()

                # Define common dynamic field mappings
                dynamic_mappings = {
                    "location": "/api/config/locations/",
                    "department": "/api/config/departments/",
                    "designation": "/api/config/designations/",
                    "role": "/api/config/roles/",
                    "employee": "/api/config/employees/list/",
                    "user": "/api/config/employees/list/",
                }

                # Check if this field should be dynamic based on its name or label
                for keyword, api_endpoint in dynamic_mappings.items():
                    if keyword in field_type_lower or keyword in field_label_lower:
                        is_dynamic = True
                        endpoint = api_endpoint
                        break

                # Extract options from additional_info
                options = field.additional_info.get("options", []) if field.additional_info else []

                # Build validation object based on data type
                validation = {}
                if field.field_type.data_type.name == "select":
                    validation = {
                        "isMultiple": field.additional_info.get("isMultiple", False) if field.additional_info else False,
                        "maxSelection": field.additional_info.get("maxSelection", "") if field.additional_info else "",
                        "minSelection": 1 if field.required else 0
                    }
                elif field.field_type.data_type.name == "text":
                    validation = {
                        "pattern": field.additional_info.get("pattern", "") if field.additional_info else "",
                        "maxLength": field.additional_info.get("maxLength", "") if field.additional_info else "",
                        "minLength": field.additional_info.get("minLength", 0) if field.additional_info else 0
                    }
                elif field.field_type.data_type.name == "number":
                    validation = {
                        "min": field.additional_info.get("min", "") if field.additional_info else "",
                        "max": field.additional_info.get("max", "") if field.additional_info else ""
                    }
                elif field.field_type.data_type.name == "file":
                    validation = {
                        "fileType": field.additional_info.get("fileType", "") if field.additional_info else "",
                        "isMultiple": field.additional_info.get("isMultiple", False) if field.additional_info else False,
                        "maxFileSize": field.additional_info.get("maxFileSize", "") if field.additional_info else ""
                    }
                elif field.field_type.data_type.name == "date":
                    validation = {
                        "startDateBeforeOrEqualEndDate": True
                    }

                field_data = {
                    "name": field_name,
                    "type": field.field_type.data_type.name,
                    "label": field.label,
                    "value": None,
                    "width": field.additional_info.get("width", "100") if field.additional_info else "100",
                    "fields": [],  # For nested fields
                    "dynamic": is_dynamic,
                    "options": options,
                    "type_id": str(field.field_type.id),
                    "position": field.order - 1,  # 0-indexed for frontend
                    "required": field.required,
                    "end_point": endpoint,
                    "dependency": field.dependency if field.dependency else {
                        "field_name": "",
                        "field_section": "",
                        "options_selected": [],
                        "cascader_selection": [],
                        "multiple_field_dependencies": []
                    },
                    "validation": validation
                }
                section_data["fields"].append(field_data)

            draft_data["sections"].append(section_data)

        # Create FormDraft with populated data
        print(f">>> Creating FormDraft with form structure...")
        FormDraft.objects.create(
            form=form,
            draft_data=draft_data
        )
        print(f">>> FormDraft created with {len(draft_data['sections'])} sections")

        print(f"\n=== Form Creation Summary ===")
        print(f"Form: {form.title} (ID: {form.id})")
        print(f"Sections created: {sections_created}")
        print(f"Fields created: {fields_created}")
        print(f"FormDraft created: Yes")

        return {
            "status": "success",
            "message": f"Form '{form_title}' created with {sections_created} sections and {fields_created} fields.",
            "form_info": {
                "id": str(form.id),
                "unique_code": form.unique_code,
                "title": form.title,
                "sections_count": sections_created,
                "fields_count": fields_created
            }
        }

    except Exception as e:
        import traceback
        print(f"\n>>> EXCEPTION in create_single_form:")
        print(f">>> Error: {str(e)}")
        print(f">>> Traceback:")
        traceback.print_exc()
        logger.exception(f"Error creating form '{form_title}': {e}")
        return {
            "status": "failed",
            "message": f"Form '{form_title}': Error - {str(e)}"
        }
