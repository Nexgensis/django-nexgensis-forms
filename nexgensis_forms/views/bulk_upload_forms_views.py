"""
Bulk Upload Views for FormsApp

Provides API endpoints for:
1. Downloading Excel template for bulk form upload
2. Uploading and processing bulk form data from Excel files
"""

import logging
import tempfile
import os
from django.http import FileResponse
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

from ..models import FormType, FieldType, DataType, Form, FormSections, FormFields
from ..services import bulk_upload_forms_services
from django.db.models import Max, Q
import json

logger = logging.getLogger(__name__)


# ===============================
# Swagger Documentation
# ===============================

# Request body for file upload
bulk_upload_request_body = openapi.Schema(
    type=openapi.TYPE_OBJECT,
    properties={
        "file": openapi.Schema(
            type=openapi.TYPE_STRING,
            format="binary",
            description="Excel file (.xlsx or .xls) to upload"
        )
    },
    required=["file"]
)

# Success example
success_example = {
    "status": "success",
    "message": "Bulk upload for forms processed successfully.",
    "total_processed": 3,
    "total_success": 3,
    "total_errors": 0,
    "successes": [
        "Form 'Safety Inspection' created with 3 sections and 12 fields.",
        "Form 'Quality Audit' created with 2 sections and 8 fields.",
    ],
    "errors": [],
    "created_forms": [
        {
            "id": "uuid-here",
            "unique_code": "FORM-00123",
            "title": "Safety Inspection",
            "sections_count": 3,
            "fields_count": 12
        }
    ]
}

# Error example
error_example = {
    "status": "failed",
    "message": "Validation failed. Please fix the errors below and re-upload the file.",
    "errors": [
        {
            "row": 5,
            "sheet": "Fields",
            "type": "Invalid Dropdown Value",
            "column": "Field Type",
            "invalid_value": "invalid_type",
            "message": "'invalid_type' is not a valid Field Type."
        }
    ],
    "total_errors": 1
}


# ===============================
# API Endpoints
# ===============================

@swagger_auto_schema(
    method="post",
    operation_description="Upload bulk form data via Excel file. The file must follow the template structure with Forms, Sections, and Fields sheets.",
    request_body=bulk_upload_request_body,
    consumes=["multipart/form-data"],
    responses={
        200: openapi.Response(
            description="Bulk upload processed successfully (or with partial success)",
            examples={"application/json": success_example}
        ),
        400: openapi.Response(
            description="Bad Request (missing file, invalid format, or validation errors)",
            examples={"application/json": error_example}
        ),
        500: openapi.Response(
            description="Internal Server Error",
            examples={"application/json": {
                "status": "failed",
                "message": "Error processing file. Please contact administrator."
            }}
        ),
    }
)
@api_view(["POST"])
@permission_classes([IsAuthenticated])
@transaction.atomic
def bulk_upload_forms(request):
    """
    Upload and process bulk form data from Excel file.

    Accepts an Excel file with three sheets:
    - Forms: Form metadata
    - Sections: Section definitions
    - Fields: Field definitions

    Validates all data before processing and returns detailed error messages.
    """
    excel_file = request.FILES.get("file")

    if not excel_file:
        return Response(
            {"status": "failed", "message": "No file uploaded"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Process the bulk upload
    result = bulk_upload_forms_services.bulk_upload_forms_service(
        excel_file, request.user
    )

    # Return appropriate status code based on result
    if result.get("status") == "failed":
        return Response(result, status=status.HTTP_400_BAD_REQUEST)

    return Response(result, status=status.HTTP_200_OK)


@swagger_auto_schema(
    method="get",
    operation_description="Download Excel template for bulk form upload. The template includes three sheets with dropdowns and validation.",
    responses={
        200: openapi.Response(
            description="Excel template file",
            content={
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            }
        ),
    }
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_forms_template(request):
    """
    Generate and download Excel template for bulk form upload.

    Template includes:
    - Sheet 1: Forms (form metadata)
    - Sheet 2: Sections (section definitions)
    - Sheet 3: Fields (field definitions)
    - Sheet 4: References (hidden sheet with dropdown data)

    All sheets include:
    - Title row (merged, styled)
    - Instruction row (merged, styled)
    - Header row (colored, bold)
    - Data validation dropdowns
    """
    try:
        # Fetch dropdown data from database (only active records)
        # FormType uses effective_end_date for soft delete
        form_types = list(
            FormType.objects.filter(effective_end_date__isnull=True)
            .values_list("name", flat=True)
            .order_by("name")
        )

        # FieldType uses is_deleted for soft delete
        field_types = list(
            FieldType.objects.filter(is_deleted=False)
            .values_list("name", flat=True)
            .order_by("name")
        )

        # DataType uses is_deleted for soft delete
        data_types = list(
            DataType.objects.filter(is_deleted=False)
            .values_list("name", flat=True)
            .order_by("name")
        )

        # Create workbook
        wb = Workbook()

        # Define common variables for data validation
        num_rows = 100  # Number of rows available for user input
        start_row = 4   # First editable row (after title, note, and header rows)

        # =============================
        # SHEET 1: FORMS
        # =============================
        ws_forms = wb.active
        ws_forms.title = "Forms"

        # Title row
        ws_forms.merge_cells("A1:D1")
        title_cell = ws_forms["A1"]
        title_cell.value = "Bulk Upload Forms - Form Metadata"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_forms.merge_cells("A2:D2")
        note_cell = ws_forms["A2"]
        note_cell.value = (
            "NOTE: Do not modify header names. Select Form Type from dropdown. "
            "Form Title must be unique."
        )
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row (row 3)
        forms_headers = ["Form Title", "Form Type", "Description", "Is Completed"]
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for col_num, header in enumerate(forms_headers, 1):
            cell = ws_forms.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws_forms.column_dimensions["A"].width = 30
        ws_forms.column_dimensions["B"].width = 25
        ws_forms.column_dimensions["C"].width = 40
        ws_forms.column_dimensions["D"].width = 15

        # Add conditional formatting to highlight duplicate Form Titles
        dxf = DifferentialStyle(
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            font=Font(color="9C0006", bold=True)
        )
        rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=False)
        ws_forms.conditional_formatting.add("A4:A103", rule)

        # =============================
        # SHEET 2: SECTIONS
        # =============================
        ws_sections = wb.create_sheet("Sections")

        # Title row
        ws_sections.merge_cells("A1:H1")
        title_cell = ws_sections["A1"]
        title_cell.value = "Bulk Upload Forms - Sections"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_sections.merge_cells("A2:H2")
        note_cell = ws_sections["A2"]
        note_cell.value = (
            "NOTE: Select Form Title from dropdown. Section Order must be unique within each form. "
            "For simple dependencies, use the three dependency columns (Section, Field, Option) instead of the JSON Dependency column."
        )
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        sections_headers = [
            "Form Title", "Section Name", "Section Description", "Section Order",
            "Dependency Section", "Dependency Field", "Dependency Option", "Dependency (JSON)"
        ]

        for col_num, header in enumerate(sections_headers, 1):
            cell = ws_sections.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws_sections.column_dimensions["A"].width = 30
        ws_sections.column_dimensions["B"].width = 30
        ws_sections.column_dimensions["C"].width = 40
        ws_sections.column_dimensions["D"].width = 15
        ws_sections.column_dimensions["E"].width = 25  # Dependency Section
        ws_sections.column_dimensions["F"].width = 25  # Dependency Field
        ws_sections.column_dimensions["G"].width = 25  # Dependency Option
        ws_sections.column_dimensions["H"].width = 30  # Dependency (JSON)

        # Add dynamic dropdown for Form Title (uses OFFSET formula for dynamic range)
        # This formula counts non-empty cells in Forms!A4:A103 and creates a dynamic list
        form_title_sections_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Forms!$A$4,0,0,COUNTA(Forms!$A$4:$A$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Form Title",
            error="Please select a Form Title from the Forms sheet"
        )
        ws_sections.add_data_validation(form_title_sections_dv)
        form_title_col_sections = sections_headers.index("Form Title") + 1
        form_title_sections_dv.add(
            f"{get_column_letter(form_title_col_sections)}{start_row}:"
            f"{get_column_letter(form_title_col_sections)}{start_row + num_rows}"
        )

        # =============================
        # SHEET 3: FIELDS
        # =============================
        ws_fields = wb.create_sheet("Fields")

        # Title row
        ws_fields.merge_cells("A1:P1")
        title_cell = ws_fields["A1"]
        title_cell.value = "Bulk Upload Forms - Fields"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_fields.merge_cells("A2:P2")
        note_cell = ws_fields["A2"]
        note_cell.value = (
            "NOTE: Select Form Title, Section Name, Field Type, Data Type, and Width from dropdowns. "
            "Options format: [\"Option1\",\"Option2\"]. "
            "For field dependency, use the three columns (Field Dep Section/Field/Option) or JSON Dependency column."
        )
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        fields_headers = [
            "Form Title", "Section Name", "Field Label",
            "Field Type", "Data Type", "Required", "Field Order",
            "Width", "Options", "Validation", "Parent Field",
            "Field Dep Section", "Field Dep Field", "Field Dep Option",
            "Dependency", "Additional Info"
        ]

        for col_num, header in enumerate(fields_headers, 1):
            cell = ws_fields.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        column_widths = {
            "A": 30, "B": 25, "C": 25,
            "D": 20, "E": 20, "F": 12, "G": 12,
            "H": 15, "I": 30, "J": 40, "K": 25,
            "L": 20, "M": 20, "N": 20,
            "O": 35, "P": 30
        }
        for col_letter, width in column_widths.items():
            ws_fields.column_dimensions[col_letter].width = width

        # Add dynamic dropdown for Form Title (uses OFFSET formula for dynamic range)
        # This formula counts non-empty cells in Forms!A4:A103 and creates a dynamic list
        form_title_fields_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Forms!$A$4,0,0,COUNTA(Forms!$A$4:$A$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Form Title",
            error="Please select a Form Title from the Forms sheet"
        )
        ws_fields.add_data_validation(form_title_fields_dv)
        form_title_col_fields = fields_headers.index("Form Title") + 1
        form_title_fields_dv.add(
            f"{get_column_letter(form_title_col_fields)}{start_row}:"
            f"{get_column_letter(form_title_col_fields)}{start_row + num_rows}"
        )

        # Add dynamic dropdown for Section Name (uses OFFSET formula for dynamic range)
        # This formula counts non-empty cells in Sections!B4:B103 and creates a dynamic list
        section_name_fields_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Sections!$B$4,0,0,COUNTA(Sections!$B$4:$B$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Section Name",
            error="Please select a Section Name from the Sections sheet"
        )
        ws_fields.add_data_validation(section_name_fields_dv)
        section_name_col_fields = fields_headers.index("Section Name") + 1
        section_name_fields_dv.add(
            f"{get_column_letter(section_name_col_fields)}{start_row}:"
            f"{get_column_letter(section_name_col_fields)}{start_row + num_rows}"
        )

        # =============================
        # SHEET 4: VALIDATION RULES (Visible - Reference for users)
        # =============================
        ws_validation = wb.create_sheet("Validation Rules")

        # Title row
        ws_validation.merge_cells("A1:D1")
        title_cell = ws_validation["A1"]
        title_cell.value = "Validation Rules by Data Type"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_validation.merge_cells("A2:D2")
        note_cell = ws_validation["A2"]
        note_cell.value = (
            "Use the validation keys below in JSON format in the Validation column. "
            "Example: {\"minLength\": 5, \"maxLength\": 100}"
        )
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        validation_headers = ["Data Type", "Validation Keys", "Description", "Example"]
        for col_num, header in enumerate(validation_headers, 1):
            cell = ws_validation.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch validation rules from DataType model
        data_type_rules = DataType.objects.filter(is_deleted=False).values('name', 'validation_rules')

        # Populate validation rules data
        row_num = 4
        for dt in data_type_rules:
            data_type_name = dt['name']
            rules = dt['validation_rules'] or []

            ws_validation[f"A{row_num}"] = data_type_name
            ws_validation[f"B{row_num}"] = ", ".join(rules) if rules else "None"

            # Add descriptions and examples based on data type
            descriptions = {
                "text": "minLength: min chars, maxLength: max chars, pattern: regex",
                "textarea": "minLength: min chars, maxLength: max chars, pattern: regex",
                "number": "min: min value, max: max value, isInteger: true/false, isPositive: true/false",
                "date": "minDate: YYYY-MM-DD, maxDate: YYYY-MM-DD",
                "date_range": "startDateBeforeOrEqualEndDate: true/false",
                "time": "minTime: HH:MM, maxTime: HH:MM",
                "time_range": "startTimeBeforeOrEqualEndTime: true/false",
                "select": "minSelection: min items, maxSelection: max items, isMultiple: true/false",
                "checkbox": "minSelection: min items, maxSelection: max items, isMultiple: true/false",
                "file": "fileType: extensions, maxFileSize: MB, isMultiple: true/false",
                "image": "maxFileSize: MB, resolution: WxH, aspectRatio: W:H, isMultiple: true/false",
                "password": "minLength: min chars, maxLength: max chars, containsSpecialChar: true/false, strengthCheck: true/false",
                "range": "min: min value, max: max value, step: increment",
                "signature": "maxSize: KB, maxDimensions: WxH",
                "richtext": "minContentLength: min chars, maxContentLength: max chars, disallowedTags: array",
            }

            examples = {
                "text": '{"minLength": 5, "maxLength": 100}',
                "textarea": '{"minLength": 10, "maxLength": 500}',
                "number": '{"min": 0, "max": 100, "isInteger": true}',
                "date": '{"minDate": "2024-01-01", "maxDate": "2025-12-31"}',
                "date_range": '{"startDateBeforeOrEqualEndDate": true}',
                "time": '{"minTime": "09:00", "maxTime": "18:00"}',
                "time_range": '{"startTimeBeforeOrEqualEndTime": true}',
                "select": '{"isMultiple": true, "minSelection": 1, "maxSelection": 3}',
                "checkbox": '{"isMultiple": true, "maxSelection": 5}',
                "file": '{"fileType": ".pdf,.doc", "maxFileSize": 10, "isMultiple": false}',
                "image": '{"maxFileSize": 5, "isMultiple": true}',
                "password": '{"minLength": 8, "containsSpecialChar": true}',
                "range": '{"min": 0, "max": 100, "step": 5}',
                "signature": '{"maxSize": 500}',
                "richtext": '{"maxContentLength": 5000}',
            }

            ws_validation[f"C{row_num}"] = descriptions.get(data_type_name, "")
            ws_validation[f"D{row_num}"] = examples.get(data_type_name, "")
            row_num += 1

        # Set column widths
        ws_validation.column_dimensions["A"].width = 15
        ws_validation.column_dimensions["B"].width = 50
        ws_validation.column_dimensions["C"].width = 60
        ws_validation.column_dimensions["D"].width = 50

        # =============================
        # SHEET 5: REFERENCES (Hidden)
        # =============================
        ws_ref = wb.create_sheet("References")

        # Form Types
        ws_ref["A1"] = "Form Types"
        ws_ref["A1"].font = Font(bold=True)
        for i, name in enumerate(form_types, start=2):
            ws_ref[f"A{i}"] = name

        # Field Types
        ws_ref["B1"] = "Field Types"
        ws_ref["B1"].font = Font(bold=True)
        for i, name in enumerate(field_types, start=2):
            ws_ref[f"B{i}"] = name

        # Data Types
        ws_ref["C1"] = "Data Types"
        ws_ref["C1"].font = Font(bold=True)
        for i, name in enumerate(data_types, start=2):
            ws_ref[f"C{i}"] = name

        # Boolean values
        ws_ref["D1"] = "Boolean"
        ws_ref["D1"].font = Font(bold=True)
        ws_ref["D2"] = "TRUE"
        ws_ref["D3"] = "FALSE"

        # Width options
        ws_ref["E1"] = "Width"
        ws_ref["E1"].font = Font(bold=True)
        width_options = [
            "25% (1/4)",
            "33% (1/3)",
            "50% (1/2)",
            "66% (2/3)",
            "75% (3/4)",
            "100% (Full)"
        ]
        for i, width_value in enumerate(width_options, start=2):
            ws_ref[f"E{i}"] = width_value

        # Hide the reference sheet
        ws_ref.sheet_state = "hidden"

        # =============================
        # DATA VALIDATION (Dropdowns)
        # =============================
        # Forms sheet - Form Type dropdown
        if form_types:
            max_ref_row = len(form_types) + 1
            form_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$A$2:$A${max_ref_row}",
                allow_blank=False
            )
            ws_forms.add_data_validation(form_type_dv)
            form_type_col = forms_headers.index("Form Type") + 1
            form_type_dv.add(
                f"{get_column_letter(form_type_col)}{start_row}:"
                f"{get_column_letter(form_type_col)}{start_row + num_rows}"
            )

        # Forms sheet - Is Completed dropdown
        is_completed_dv = DataValidation(
            type="list",
            formula1="=References!$D$2:$D$3",
            allow_blank=True
        )
        ws_forms.add_data_validation(is_completed_dv)
        is_completed_col = forms_headers.index("Is Completed") + 1
        is_completed_dv.add(
            f"{get_column_letter(is_completed_col)}{start_row}:"
            f"{get_column_letter(is_completed_col)}{start_row + num_rows}"
        )

        # Fields sheet - Field Type dropdown
        if field_types:
            max_ref_row = len(field_types) + 1
            field_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$B$2:$B${max_ref_row}",
                allow_blank=False
            )
            ws_fields.add_data_validation(field_type_dv)
            field_type_col = fields_headers.index("Field Type") + 1
            field_type_dv.add(
                f"{get_column_letter(field_type_col)}{start_row}:"
                f"{get_column_letter(field_type_col)}{start_row + num_rows}"
            )

        # Fields sheet - Data Type dropdown
        if data_types:
            max_ref_row = len(data_types) + 1
            data_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$C$2:$C${max_ref_row}",
                allow_blank=False
            )
            ws_fields.add_data_validation(data_type_dv)
            data_type_col = fields_headers.index("Data Type") + 1
            data_type_dv.add(
                f"{get_column_letter(data_type_col)}{start_row}:"
                f"{get_column_letter(data_type_col)}{start_row + num_rows}"
            )

        # Fields sheet - Required dropdown
        required_dv = DataValidation(
            type="list",
            formula1="=References!$D$2:$D$3",
            allow_blank=True
        )
        ws_fields.add_data_validation(required_dv)
        required_col = fields_headers.index("Required") + 1
        required_dv.add(
            f"{get_column_letter(required_col)}{start_row}:"
            f"{get_column_letter(required_col)}{start_row + num_rows}"
        )

        # Fields sheet - Width dropdown
        width_dv = DataValidation(
            type="list",
            formula1="=References!$E$2:$E$7",
            allow_blank=True
        )
        ws_fields.add_data_validation(width_dv)
        width_col = fields_headers.index("Width") + 1
        width_dv.add(
            f"{get_column_letter(width_col)}{start_row}:"
            f"{get_column_letter(width_col)}{start_row + num_rows}"
        )

        # =============================
        # ADD SAMPLE DATA (Optional)
        # =============================
        # You can uncomment this section to add sample rows

        # Sample form
        # ws_forms["A4"] = "Safety Inspection Checklist"
        # ws_forms["B4"] = form_types[0] if form_types else ""
        # ws_forms["C4"] = "Monthly safety inspection form"
        # ws_forms["D4"] = "TRUE"

        # Sample section
        # ws_sections["A4"] = "Safety Inspection Checklist"
        # ws_sections["B4"] = "General Information"
        # ws_sections["C4"] = "Basic inspection details"
        # ws_sections["D4"] = 1

        # Sample field
        # ws_fields["A4"] = "Safety Inspection Checklist"
        # ws_fields["B4"] = "General Information"
        # ws_fields["C4"] = "Inspector Name"
        # ws_fields["D4"] = "inspector_name"
        # ws_fields["E4"] = field_types[0] if field_types else ""
        # ws_fields["F4"] = data_types[0] if data_types else ""
        # ws_fields["G4"] = "TRUE"
        # ws_fields["H4"] = 1

        # =============================
        # SAVE AND RETURN FILE
        # =============================
        # Save workbook to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        # Open for reading
        file_handle = open(temp_file.name, 'rb')

        # Return downloadable file
        response = FileResponse(
            file_handle,
            as_attachment=True,
            filename="Forms_Bulk_Upload_Template.xlsx"
        )

        # Cleanup on close
        def cleanup_file():
            try:
                file_handle.close()
                os.remove(temp_file.name)
            except OSError:
                pass

        response.close = cleanup_file

        return response

    except Exception as e:
        logger.exception(f"Error generating template: {e}")
        return Response(
            {"status": "failed", "message": f"Error generating template: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@swagger_auto_schema(
    method="get",
    operation_description="Export all forms (latest versions only), sections, and fields from the database to an Excel file. "
                          "This file can be downloaded and uploaded to another deployment site.",
    responses={
        200: openapi.Response(
            description="Excel file with all forms data",
            schema=openapi.Schema(
                type=openapi.TYPE_FILE
            )
        ),
        500: openapi.Response(
            description="Internal Server Error",
            examples={"application/json": {"status": "failed", "message": "Error exporting forms"}}
        ),
    }
)
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_forms_data(request):
    """
    Export all forms (latest versions only), sections, and fields from database to Excel format.

    The generated Excel file matches the bulk upload template format and contains:
    - Forms sheet: Latest version of all forms with their metadata
    - Sections sheet: All sections with their form associations
    - Fields sheet: All fields with their complete configuration
    - References sheet: Dropdown options (hidden)

    Only the latest version of each form is exported (based on root_form and version number).
    This allows easy migration of forms between deployments.
    """
    from django.db.models import Prefetch

    try:
        logger.info("Starting forms data export...")

        # Fetch all active forms (not deleted)
        # Form model uses effective_end_date for soft delete
        # Get only the latest version of each form by grouping by root_form

        # First, get the maximum version for each root_form
        latest_versions = Form.objects.filter(
            effective_end_date__isnull=True
        ).values('root_form').annotate(
            max_version=Max('version')
        )

        # Build a list of (root_form, max_version) tuples
        latest_version_filters = Q()
        for item in latest_versions:
            latest_version_filters |= Q(
                root_form=item['root_form'],
                version=item['max_version']
            )

        # Prefetch sections with their fields to avoid N+1 queries
        # This fetches all related data in just a few queries instead of hundreds
        # Note: Using default related names (formsections_set, formfields_set)
        fields_prefetch = Prefetch(
            'formfields_set',
            queryset=FormFields.objects.filter(is_deleted=False)
                .select_related('field_type', 'field_type__data_type', 'parent_field')
                .order_by('order', 'created_on')
        )

        sections_prefetch = Prefetch(
            'formsections_set',
            queryset=FormSections.objects.filter(is_deleted=False)
                .prefetch_related(fields_prefetch)
                .order_by('order', 'created_on')
        )

        # Fetch only the latest versions with all related data prefetched
        forms = Form.objects.filter(
            effective_end_date__isnull=True
        ).filter(
            latest_version_filters
        ).select_related('form_type').prefetch_related(sections_prefetch).order_by('title')

        if not forms.exists():
            return Response(
                {"status": "failed", "message": "No forms found to export."},
                status=status.HTTP_404_NOT_FOUND
            )

        logger.info(f"Found {forms.count()} forms (latest versions only) to export")

        # Fetch reference data for dropdowns
        form_types = list(
            FormType.objects.filter(effective_end_date__isnull=True)
            .values_list("name", flat=True)
            .order_by("name")
        )
        field_types = list(
            FieldType.objects.filter(is_deleted=False)
            .values_list("name", flat=True)
            .order_by("name")
        )
        data_types = list(
            DataType.objects.filter(is_deleted=False)
            .values_list("name", flat=True)
            .order_by("name")
        )

        # Create workbook
        wb = Workbook()

        # Define common variables
        num_rows = 100
        start_row = 4

        # =============================
        # SHEET 1: FORMS
        # =============================
        ws_forms = wb.active
        ws_forms.title = "Forms"

        # Title row
        ws_forms.merge_cells("A1:D1")
        title_cell = ws_forms["A1"]
        title_cell.value = "Bulk Upload Forms - Data Export"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_forms.merge_cells("A2:D2")
        note_cell = ws_forms["A2"]
        note_cell.value = "NOTE: This file contains exported data (latest versions only) from the database. You can upload it to another deployment."
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        forms_headers = ["Form Title", "Form Type", "Description", "Is Completed"]
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for col_num, header in enumerate(forms_headers, 1):
            cell = ws_forms.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws_forms.column_dimensions["A"].width = 30
        ws_forms.column_dimensions["B"].width = 25
        ws_forms.column_dimensions["C"].width = 40
        ws_forms.column_dimensions["D"].width = 15

        # Populate forms data
        row_num = start_row
        for form in forms:
            ws_forms[f"A{row_num}"] = form.title
            ws_forms[f"B{row_num}"] = form.form_type.name if form.form_type else ""
            ws_forms[f"C{row_num}"] = form.description or ""
            ws_forms[f"D{row_num}"] = "TRUE" if form.is_completed else "FALSE"
            row_num += 1

        logger.info(f"Exported {row_num - start_row} forms to Forms sheet")

        # Add conditional formatting for duplicates
        dxf = DifferentialStyle(
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            font=Font(color="9C0006", bold=True)
        )
        rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=False)
        ws_forms.conditional_formatting.add("A4:A103", rule)

        # =============================
        # SHEET 2: SECTIONS
        # =============================
        ws_sections = wb.create_sheet("Sections")

        # Title row
        ws_sections.merge_cells("A1:E1")
        title_cell = ws_sections["A1"]
        title_cell.value = "Form Sections"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_sections.merge_cells("A2:E2")
        note_cell = ws_sections["A2"]
        note_cell.value = "NOTE: Select Form Title from dropdown. Section Order must be unique within each form."
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        sections_headers = [
            "Form Title", "Section Name", "Section Description", "Section Order", "Dependency (JSON)"
        ]

        for col_num, header in enumerate(sections_headers, 1):
            cell = ws_sections.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws_sections.column_dimensions["A"].width = 30
        ws_sections.column_dimensions["B"].width = 30
        ws_sections.column_dimensions["C"].width = 40
        ws_sections.column_dimensions["D"].width = 15
        ws_sections.column_dimensions["E"].width = 40

        # Populate sections data (using prefetched data - no additional queries)
        row_num = start_row
        sections_count = 0
        for form in forms:
            # Use prefetched sections instead of querying again
            sections = list(form.formsections_set.all())

            # Enumerate sections to assign sequential order numbers (1, 2, 3, ...)
            for idx, section in enumerate(sections, start=1):
                ws_sections[f"A{row_num}"] = form.title
                ws_sections[f"B{row_num}"] = section.name
                ws_sections[f"C{row_num}"] = section.description or ""
                ws_sections[f"D{row_num}"] = idx  # Use enumerated index instead of section.order
                ws_sections[f"E{row_num}"] = json.dumps(section.dependency) if section.dependency else ""
                row_num += 1
                sections_count += 1

        logger.info(f"Exported {sections_count} sections to Sections sheet")

        # =============================
        # SHEET 3: FIELDS
        # =============================
        ws_fields = wb.create_sheet("Fields")

        # Title row
        ws_fields.merge_cells("A1:L1")
        title_cell = ws_fields["A1"]
        title_cell.value = "Form Fields"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Note row
        ws_fields.merge_cells("A2:L2")
        note_cell = ws_fields["A2"]
        note_cell.value = (
            "NOTE: Select Form Title, Section Name, Field Type, Data Type, and Width from dropdowns. "
            "Options should be JSON array like: [\"Option1\",\"Option2\"]. Field Name will be auto-generated."
        )
        note_cell.font = Font(size=10, italic=True, color="FF0000")
        note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Header row
        fields_headers = [
            "Form Title", "Section Name", "Field Label",
            "Field Type", "Data Type", "Required", "Field Order",
            "Width", "Options", "Parent Field", "Dependency", "Additional Info"
        ]

        for col_num, header in enumerate(fields_headers, 1):
            cell = ws_fields.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        column_widths = {
            "A": 30, "B": 25, "C": 25,
            "D": 20, "E": 20, "F": 12, "G": 12,
            "H": 15, "I": 30, "J": 25, "K": 30, "L": 30
        }
        for col_letter, width in column_widths.items():
            ws_fields.column_dimensions[col_letter].width = width

        # Populate fields data (using prefetched data - no additional queries)
        row_num = start_row
        fields_count = 0

        # Width display mapping (defined once, reused)
        width_display_map = {
            "25": "25% (1/4)",
            "33": "33% (1/3)",
            "50": "50% (1/2)",
            "66": "66% (2/3)",
            "75": "75% (3/4)",
            "100": "100% (Full)"
        }

        for form in forms:
            # Use prefetched sections instead of querying again
            for section in form.formsections_set.all():
                # Get all fields from prefetched data and separate parent/child fields
                all_fields = list(section.formfields_set.all())
                parent_fields = [f for f in all_fields if f.parent_field is None]
                child_fields_by_parent = {}
                for f in all_fields:
                    if f.parent_field_id:
                        if f.parent_field_id not in child_fields_by_parent:
                            child_fields_by_parent[f.parent_field_id] = []
                        child_fields_by_parent[f.parent_field_id].append(f)

                # Use a counter for field order within each section
                field_order_counter = 1

                for field in parent_fields:
                    ws_fields[f"A{row_num}"] = form.title
                    ws_fields[f"B{row_num}"] = section.name
                    ws_fields[f"C{row_num}"] = field.label
                    ws_fields[f"D{row_num}"] = field.field_type.name if field.field_type else ""
                    ws_fields[f"E{row_num}"] = field.field_type.data_type.name if field.field_type and field.field_type.data_type else ""
                    ws_fields[f"F{row_num}"] = "TRUE" if field.required else "FALSE"
                    ws_fields[f"G{row_num}"] = field_order_counter
                    field_order_counter += 1

                    # Extract width from additional_info
                    width = field.additional_info.get("width", "100") if field.additional_info else "100"
                    ws_fields[f"H{row_num}"] = width_display_map.get(str(width), f"{width}%")

                    # Extract options from additional_info
                    options = field.additional_info.get("options", []) if field.additional_info else []
                    ws_fields[f"I{row_num}"] = json.dumps(options) if options else ""

                    # Parent field is empty for parent fields
                    ws_fields[f"J{row_num}"] = ""

                    # Dependency
                    ws_fields[f"K{row_num}"] = json.dumps(field.dependency) if field.dependency else ""

                    # Additional Info (excluding width and options)
                    if field.additional_info:
                        filtered_info = {k: v for k, v in field.additional_info.items()
                                       if k not in ["width", "options"]}
                        ws_fields[f"L{row_num}"] = json.dumps(filtered_info) if filtered_info else ""
                    else:
                        ws_fields[f"L{row_num}"] = ""

                    row_num += 1
                    fields_count += 1

                    # Handle sub-fields from prefetched data (no additional query)
                    sub_fields = child_fields_by_parent.get(field.id, [])

                    for sub_field in sub_fields:
                        ws_fields[f"A{row_num}"] = form.title
                        ws_fields[f"B{row_num}"] = section.name
                        ws_fields[f"C{row_num}"] = sub_field.label
                        ws_fields[f"D{row_num}"] = sub_field.field_type.name if sub_field.field_type else ""
                        ws_fields[f"E{row_num}"] = sub_field.field_type.data_type.name if sub_field.field_type and sub_field.field_type.data_type else ""
                        ws_fields[f"F{row_num}"] = "TRUE" if sub_field.required else "FALSE"
                        ws_fields[f"G{row_num}"] = field_order_counter
                        field_order_counter += 1

                        # Width
                        width = sub_field.additional_info.get("width", "100") if sub_field.additional_info else "100"
                        ws_fields[f"H{row_num}"] = width_display_map.get(str(width), f"{width}%")

                        # Options
                        options = sub_field.additional_info.get("options", []) if sub_field.additional_info else []
                        ws_fields[f"I{row_num}"] = json.dumps(options) if options else ""

                        # Parent field label
                        ws_fields[f"J{row_num}"] = field.label

                        # Dependency
                        ws_fields[f"K{row_num}"] = json.dumps(sub_field.dependency) if sub_field.dependency else ""

                        # Additional Info
                        if sub_field.additional_info:
                            filtered_info = {k: v for k, v in sub_field.additional_info.items()
                                           if k not in ["width", "options"]}
                            ws_fields[f"L{row_num}"] = json.dumps(filtered_info) if filtered_info else ""
                        else:
                            ws_fields[f"L{row_num}"] = ""

                        row_num += 1
                        fields_count += 1

        logger.info(f"Exported {fields_count} fields to Fields sheet")

        # =============================
        # SHEET 4: REFERENCES (Hidden)
        # =============================
        ws_ref = wb.create_sheet("References")

        # Form Types
        ws_ref["A1"] = "Form Types"
        ws_ref["A1"].font = Font(bold=True)
        for i, form_type in enumerate(form_types, start=2):
            ws_ref[f"A{i}"] = form_type

        # Field Types
        ws_ref["B1"] = "Field Types"
        ws_ref["B1"].font = Font(bold=True)
        for i, field_type in enumerate(field_types, start=2):
            ws_ref[f"B{i}"] = field_type

        # Data Types
        ws_ref["C1"] = "Data Types"
        ws_ref["C1"].font = Font(bold=True)
        for i, data_type in enumerate(data_types, start=2):
            ws_ref[f"C{i}"] = data_type

        # Required options
        ws_ref["D1"] = "Required"
        ws_ref["D1"].font = Font(bold=True)
        ws_ref["D2"] = "TRUE"
        ws_ref["D3"] = "FALSE"

        # Width options
        ws_ref["E1"] = "Width"
        ws_ref["E1"].font = Font(bold=True)
        width_options = ["25% (1/4)", "33% (1/3)", "50% (1/2)", "66% (2/3)", "75% (3/4)", "100% (Full)"]
        for i, width_value in enumerate(width_options, start=2):
            ws_ref[f"E{i}"] = width_value

        # Hide the reference sheet
        ws_ref.sheet_state = "hidden"

        # =============================
        # DATA VALIDATION (Dropdowns)
        # =============================
        # Forms sheet - Form Type dropdown
        if form_types:
            max_ref_row = len(form_types) + 1
            form_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$A$2:$A${max_ref_row}",
                allow_blank=False
            )
            ws_forms.add_data_validation(form_type_dv)
            form_type_col = forms_headers.index("Form Type") + 1
            form_type_dv.add(
                f"{get_column_letter(form_type_col)}{start_row}:"
                f"{get_column_letter(form_type_col)}{start_row + num_rows}"
            )

        # Sections sheet - Form Title dropdown
        form_title_sections_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Forms!$A$4,0,0,COUNTA(Forms!$A$4:$A$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Form Title",
            error="Please select a Form Title from the Forms sheet"
        )
        ws_sections.add_data_validation(form_title_sections_dv)
        form_title_col_sections = sections_headers.index("Form Title") + 1
        form_title_sections_dv.add(
            f"{get_column_letter(form_title_col_sections)}{start_row}:"
            f"{get_column_letter(form_title_col_sections)}{start_row + num_rows}"
        )

        # Fields sheet - Form Title dropdown
        form_title_fields_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Forms!$A$4,0,0,COUNTA(Forms!$A$4:$A$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Form Title",
            error="Please select a Form Title from the Forms sheet"
        )
        ws_fields.add_data_validation(form_title_fields_dv)
        form_title_col_fields = fields_headers.index("Form Title") + 1
        form_title_fields_dv.add(
            f"{get_column_letter(form_title_col_fields)}{start_row}:"
            f"{get_column_letter(form_title_col_fields)}{start_row + num_rows}"
        )

        # Fields sheet - Section Name dropdown
        section_name_fields_dv = DataValidation(
            type="list",
            formula1="=OFFSET(Sections!$B$4,0,0,COUNTA(Sections!$B$4:$B$103),1)",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid Section Name",
            error="Please select a Section Name from the Sections sheet"
        )
        ws_fields.add_data_validation(section_name_fields_dv)
        section_name_col_fields = fields_headers.index("Section Name") + 1
        section_name_fields_dv.add(
            f"{get_column_letter(section_name_col_fields)}{start_row}:"
            f"{get_column_letter(section_name_col_fields)}{start_row + num_rows}"
        )

        # Fields sheet - Field Type dropdown
        if field_types:
            max_ref_row = len(field_types) + 1
            field_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$B$2:$B${max_ref_row}",
                allow_blank=False
            )
            ws_fields.add_data_validation(field_type_dv)
            field_type_col = fields_headers.index("Field Type") + 1
            field_type_dv.add(
                f"{get_column_letter(field_type_col)}{start_row}:"
                f"{get_column_letter(field_type_col)}{start_row + num_rows}"
            )

        # Fields sheet - Data Type dropdown
        if data_types:
            max_ref_row = len(data_types) + 1
            data_type_dv = DataValidation(
                type="list",
                formula1=f"=References!$C$2:$C${max_ref_row}",
                allow_blank=False
            )
            ws_fields.add_data_validation(data_type_dv)
            data_type_col = fields_headers.index("Data Type") + 1
            data_type_dv.add(
                f"{get_column_letter(data_type_col)}{start_row}:"
                f"{get_column_letter(data_type_col)}{start_row + num_rows}"
            )

        # Fields sheet - Required dropdown
        required_dv = DataValidation(
            type="list",
            formula1="=References!$D$2:$D$3",
            allow_blank=True
        )
        ws_fields.add_data_validation(required_dv)
        required_col = fields_headers.index("Required") + 1
        required_dv.add(
            f"{get_column_letter(required_col)}{start_row}:"
            f"{get_column_letter(required_col)}{start_row + num_rows}"
        )

        # Fields sheet - Width dropdown
        width_dv = DataValidation(
            type="list",
            formula1="=References!$E$2:$E$7",
            allow_blank=True
        )
        ws_fields.add_data_validation(width_dv)
        width_col = fields_headers.index("Width") + 1
        width_dv.add(
            f"{get_column_letter(width_col)}{start_row}:"
            f"{get_column_letter(width_col)}{start_row + num_rows}"
        )

        # =============================
        # SAVE AND RETURN FILE
        # =============================
        # Save workbook to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        # Open for reading
        file_handle = open(temp_file.name, 'rb')

        # Return downloadable file with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Forms_Export_{timestamp}.xlsx"

        response = FileResponse(
            file_handle,
            as_attachment=True,
            filename=filename
        )

        # Cleanup on close
        def cleanup_file():
            try:
                file_handle.close()
                os.remove(temp_file.name)
            except OSError:
                pass

        response.close = cleanup_file

        logger.info(f"Successfully exported {forms.count()} forms, {sections_count} sections, {fields_count} fields")

        return response

    except Exception as e:
        logger.exception(f"Error exporting forms data: {e}")
        return Response(
            {"status": "failed", "message": f"Error exporting forms data: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
